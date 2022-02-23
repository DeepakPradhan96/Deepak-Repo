import openpyxl

path="C:\Driver\selenium test\Ts.xlsx"

workbook=openpyxl.load_workbook(path)
sheet= workbook.active
#workbook.get_sheet_by_name("sheet1")

rows=sheet.max_row
colm= sheet.max_column
print(rows)
print(colm)